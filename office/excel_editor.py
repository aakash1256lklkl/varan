"""
Varan Excel editor — create, edit, read and summarize .xlsx files
built on openpyxl. Supports cell writing, formulas, styling and charts.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from . import excel_live as _live


class ExcelEditor:
    extension = ".xlsx"

    # ------------------------------------------------------------------
    # Creation
    # ------------------------------------------------------------------
    def create_workbook(self, path: str | Path, sheet: str = "Sheet1",
                        data: Optional[list[dict]] = None) -> str:
        """Create a new workbook.

        data: list of dicts, one per row: {"cells": [...], "bold": bool}
        (first row is treated as header if header=True)
        """
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet

        for row in (data or []):
            cells = row.get("cells", []) if isinstance(row, dict) else row
            ws.append(cells)

        wb.save(str(path))
        return str(path)

    # ------------------------------------------------------------------
    # Editing
    # ------------------------------------------------------------------
    def edit_workbook(self, path: str | Path, sheet: str = None,
                      writes: Optional[list[dict]] = None,
                      formulas: Optional[list[dict]] = None,
                      add_chart: Optional[dict] = None,
                      new_sheet: Optional[str] = None,
                      delete_sheet: Optional[str] = None,
                      rows: Optional[list[dict]] = None,
                      columns: Optional[list[dict]] = None,
                      clear: Optional[list[str]] = None,
                      styles: Optional[list[dict]] = None,
                      inplace: bool = False) -> str:
        """Edit an existing workbook. Returns path to the written file.

        By default writes a "NAME_edited.xlsx" copy; inplace=True saves back
        over the original (used when editing the selected target file).

        writes:    [{"cell": "A1", "value": ...}]
        formulas:  [{"cell": "C2", "formula": "=SUM(A1:A10)"}]
        add_chart: {"type": "bar"|"line"|"pie", "title": "...",
                    "categories": "A1:A5", "data": "B1:B5"}
        new_sheet: name of a sheet to ADD.
        delete_sheet: name of a sheet to DELETE (complex remove).
        rows:      [{"action": "insert"|"delete", "at": <1-based row>}]
        columns:   [{"action": "insert"|"delete", "at": "<column letter>"}]
        clear:     [range string like "A1:C5"] — clears values+styles in that
                   rectangular range.
        styles:    [{"cell": "A1", "bold": bool, "italic": bool, "size": pt,
                     "font": "Calibri", "fill": "RRGGBB"}] — apply cell styling.
        """
        src = Path(path)
        if not src.exists():
            raise FileNotFoundError(src)
        dst = src if inplace else self._edited_copy(src)
        wb = load_workbook(str(src))

        if new_sheet and not (delete_sheet == new_sheet):
            if new_sheet not in wb.sheetnames:
                wb.create_sheet(title=new_sheet)

        if delete_sheet and delete_sheet in wb.sheetnames:
            if len(wb.sheetnames) > 1:
                del wb[delete_sheet]

        target = wb[sheet] if sheet else wb.active

        for w in (writes or []):
            cell = target[w["cell"]]
            cell.value = w["value"]

        for f in (formulas or []):
            target[f["cell"]] = f["formula"]

        # -- complex row/column/clear/style operations -------------------------
        for op in (rows or []):
            self._apply_row_op(target, op)
        for op in (columns or []):
            self._apply_col_op(target, op)
        for rng in (clear or []):
            self._apply_clear(target, str(rng))
        for spec in (styles or []):
            self._apply_style(target, spec)

        if add_chart:
            self._add_chart(target, add_chart)

        wb.save(str(dst))
        return str(dst)

    @staticmethod
    def _apply_row_op(ws, op: dict) -> None:
        action = (op.get("action") or "insert").lower()
        at = int(op.get("at", 1) or 1)
        if action == "delete":
            ws.delete_rows(at, 1)
        else:
            ws.insert_rows(at, 1)

    @staticmethod
    def _apply_col_op(ws, op: dict) -> None:
        action = (op.get("action") or "insert").lower()
        letter = str(op.get("at") or "A")
        col_num = ExcelEditor._coord(f"{letter}1")[0]
        if action == "delete":
            ws.delete_cols(col_num, 1)
        else:
            ws.insert_cols(col_num, 1)

    @staticmethod
    def _apply_clear(ws, rng: str) -> None:
        if ":" in rng:
            first, last = rng.split(":")
            c1, r1 = ExcelEditor._coord(first)
            c2, r2 = ExcelEditor._coord(last)
            for row in ws.iter_rows(min_row=r1, min_col=c1,
                                    max_row=r2, max_col=c2):
                for cell in row:
                    cell.value = None
                    cell.style = "Normal"
        else:
            cell = ws[rng]
            cell.value = None
            cell.style = "Normal"

    @staticmethod
    def _apply_style(ws, spec: dict) -> None:
        cell = ws[spec.get("cell")]
        kw = {}
        if spec.get("bold") is not None:
            kw["bold"] = bool(spec["bold"])
        if spec.get("italic") is not None:
            kw["italic"] = bool(spec["italic"])
        if spec.get("size"):
            kw["size"] = float(spec["size"])
        if spec.get("font"):
            kw["name"] = str(spec["font"])
        if kw:
            cell.font = Font(**kw)
        fill = spec.get("fill")
        if fill:
            cell.fill = PatternFill(start_color=str(fill), end_color=str(fill),
                                    fill_type="solid")

    def _add_chart(self, ws, spec: dict) -> None:
        ctype = (spec.get("type") or "bar").lower()
        title = spec.get("title", "")
        cat = spec.get("categories", "")
        data = spec.get("data", "")
        if not cat or not data:
            return

        def _ref(rng: str):
            # e.g. "A1:A5" -> (min_col, min_row, max_col, max_row)
            first, last = rng.split(":")
            fc, fr = self._coord(first)
            lc, lr = self._coord(last)
            return fc, fr, lc, lr

        if ctype == "pie":
            chart = PieChart()
        elif ctype == "line":
            chart = LineChart()
        else:
            chart = BarChart()
        chart.title = title

        lc, fr, rcc, lr = _ref(cat)
        ref_cat = Reference(ws, min_col=lc, min_row=fr, max_col=rcc, max_row=lr)
        lc2, fr2, rcc2, lr2 = _ref(data)
        ref_data = Reference(ws, min_col=lc2, min_row=fr2, max_col=rcc2, max_row=lr2)

        if ctype == "pie":
            chart.add_data(ref_data, titles_from_data=True)
            chart.set_categories(ref_cat)
        else:
            chart.add_data(ref_data, titles_from_data=True)
            chart.set_categories(ref_cat)

        ws.add_chart(chart, spec.get("anchor", "F2"))

    @staticmethod
    def _coord(cell: str):
        col_letters = "".join(c for c in cell if c.isalpha())
        row_digits = "".join(c for c in cell if c.isdigit())
        # convert column letters to number
        col_num = 0
        for ch in col_letters:
            col_num = col_num * 26 + (ord(ch.upper()) - 64)
        return col_num, int(row_digits)

    def _edited_copy(self, src: Path) -> Path:
        return src.with_name(f"{src.stem}_edited{self.extension}")

    # ------------------------------------------------------------------
    # Reading
    # ------------------------------------------------------------------
    def read_workbook(self, path: str | Path) -> dict:
        src = Path(path)
        if not src.exists():
            raise FileNotFoundError(src)
        wb = load_workbook(str(src), data_only=True)
        result = {"path": str(src), "sheets": {}}
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            result["sheets"][ws.title] = rows
        return result

    def summarize(self, path: str | Path, max_rows: int = 20) -> dict:
        data = self.read_workbook(path)
        summary = {"path": data["path"], "sheets": []}
        for name, rows in data["sheets"].items():
            summary["sheets"].append({
                "name": name,
                "rows": len(rows),
                "cols": len(rows[0]) if rows else 0,
                "preview": rows[:max_rows],
            })
        return summary

    # -- live editing via Excel COM (workbook open in Microsoft Excel) ----
    def is_open_in_excel(self, path: str | Path) -> bool:
        """Return True if the .xlsx is currently open in Microsoft Excel."""
        try:
            return bool(_live.is_open_in_excel(path))
        except Exception:  # noqa: BLE001
            return False

    def live_edit(self, path: str | Path, sheet: str | None = None,
                  writes=None, formulas=None, new_sheet: str | None = None,
                  delete_sheet: str | None = None, rows=None, columns=None,
                  clear=None, styles=None, add_chart=None) -> dict:
        """Apply the same job edit_workbook supports, but LIVE in the open
        workbook via COM. Returns {'ok','path','mode':'live-excel'}.
        """
        return _live.live_edit(
            path, sheet=sheet, writes=writes, formulas=formulas,
            new_sheet=new_sheet, delete_sheet=delete_sheet,
            rows=rows, columns=columns, clear=clear, styles=styles,
            add_chart=add_chart)

    def live_append(self, path: str | Path, text: str, sheet: str | None = None) -> dict:
        """Write text into the next empty cell of column A, live on screen."""
        return _live.append_text(path, text, sheet=sheet)
